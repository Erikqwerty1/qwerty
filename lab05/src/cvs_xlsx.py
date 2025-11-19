import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_file = Path(csv_path)
    xlsx_file = Path(xlsx_path)

    if not csv_file.exists():
        raise FileNotFoundError(f"CSV-файл не найден: {csv_file}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    col_widths = {}

    with csv_file.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
            for col_index, cell_value in enumerate(row, start=1):
                length = len(str(cell_value))
                if length > col_widths.get(col_index, 0):
                    col_widths[col_index] = length

    for col_index, width in col_widths.items():
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = max(width + 2, 8)

    wb.save(xlsx_file)


if __name__ == "__main__":
    csv_path = "lab05/data/samples/people.csv"
    xlsx_path = "lab05/data/out/out_people.xlsx"

    csv_to_xlsx(csv_path, xlsx_path)