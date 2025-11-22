import argparse
import sys
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_file = Path(csv_path)
    xlsx_file = Path(xlsx_path)

    if not csv_file.exists():
        raise FileNotFoundError(f"CSV-файл не найден: {csv_file}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    col_widths = {}

    with csv_file.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
            for col_index, cell_value in enumerate(row, start=1):
                length = len(str(cell_value))
                if length > col_widths.get(col_index, 0):
                    col_widths[col_index] = length

    for col_index, width in col_widths.items():
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].width = max(width + 2, 8)

    wb.save(xlsx_file)


def json_to_csv(json_path: str, csv_path: str) -> None:
    json_file = Path(json_path)
    csv_file = Path(csv_path)

    if not json_file.exists():
        raise FileNotFoundError(f"JSON-файл не найден: {json_file}")

    with json_file.open("r", encoding="utf-8") as file:
        data = json.load(file)

    if not isinstance(data, list) or not all(isinstance(item, dict) for item in data):
        raise ValueError("Ожидался список словарей")

    fieldnames = list(data[0].keys())

    with csv_file.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in data:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def csv_to_json(csv_path: str, json_path: str) -> None:
    csv_file = Path(csv_path)
    json_file = Path(json_path)

    if not csv_file.exists():
        raise FileNotFoundError(f"CSV-файл не найден: {csv_file}")

    with csv_file.open("r", encoding="utf-8", newline="") as file:
        reader = csv.DictReader(file)
        if not reader.fieldnames:
            raise ValueError("CSV не содержит заголовка")

        data = [dict(row) for row in reader]

    with json_file.open("w", encoding="utf-8") as f:
        json.dump(data, f)


def json2csv_command(args):
    try:
        json_to_csv(args.infile, args.out)
        print(f"Успешно конвертировано: {args.infile} -> {args.out}")
    except Exception as e:
        print(f"Ошибка при конвертации JSON в CSV: {e}")
        sys.exit(1)


def csv2json_command(args):
    try:
        csv_to_json(args.infile, args.out)
        print(f"Успешно конвертировано: {args.infile} -> {args.out}")
    except Exception as e:
        print(f"Ошибка при конвертации CSV в JSON: {e}")
        sys.exit(1)


def csv2xlsx_command(args):
    try:
        csv_to_xlsx(args.infile, args.out)
        print(f"Успешно конвертировано: {args.infile} -> {args.out}")
    except Exception as e:
        print(f"Ошибка при конвертации CSV в XLSX: {e}")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Конвертер между форматами данных")

    subparsers = parser.add_subparsers(dest="command", help="Доступные команды")

    json2csv_parser = subparsers.add_parser("json2csv", help="Конвертация JSON в CSV")
    json2csv_parser.add_argument(
        "--input", dest="infile", required=True, help="Входной JSON файл"
    )
    json2csv_parser.add_argument("--out", required=True, help="Выходной CSV файл")

    csv2json_parser = subparsers.add_parser("csv2json", help="Конвертация CSV в JSON")
    csv2json_parser.add_argument(
        "--input", dest="infile", required=True, help="Входной CSV файл"
    )
    csv2json_parser.add_argument("--out", required=True, help="Выходной JSON файл")

    csv2xlsx_parser = subparsers.add_parser("csv2xlsx", help="Конвертация CSV в XLSX")
    csv2xlsx_parser.add_argument(
        "--input", dest="infile", required=True, help="Входной CSV файл"
    )
    csv2xlsx_parser.add_argument("--out", required=True, help="Выходной XLSX файл")

    args = parser.parse_args()

    if args.command == "json2csv":
        json2csv_command(args)
    elif args.command == "csv2json":
        csv2json_command(args)
    elif args.command == "csv2xlsx":
        csv2xlsx_command(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
